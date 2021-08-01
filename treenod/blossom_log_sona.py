# -*- coding: utf-8 -*-
import openpyxl
import ast

filename = 'ActionLog_20201202.xlsx'

#엑셀 파일 로드
book = openpyxl.load_workbook(filename)

#0번째 시트 선택
sheet = book.worksheets[0]

'''
data = []

for colume in sheet.columns:
        data.append(colume[1].value)

#print(data[4])
'''

api_list = {'/building/buy' : '건설', '/user/info/save' : '저장'}
request_replace = {'level' : '레벨', 'useCoin' : '코인', 'atkItemMap' : '어택아이템', 'buyItemMap' : '구입아이템', 
'rewardItemList' : '보상리스트', 'spot' : '스팟', 'step' : '스텝', 'actionKey' : '액션키', 'timestamp' : '타임스탬프', 'dataVersion' : '데이터'}
result_replace = {'sk' : '', 'uid' : '', 'ver' : '', 'm_lv' : '', 'x_fa' : '', 'm_prg' : '', 'b_amap' : '', 'b_bmap' : '', 'b_cmpl' : '', 'e_ndbt': '',
'h_coin': '', 'h_star': '', 'i_coin': '', 'i_star': ''}

#줄바꿈 문자
z = sheet['C2'].value

api = sheet['B2'].value
if api in api_list:
    api = api_list[api]

RequestJson = sheet['E2'].value
dict_request = ast.literal_eval(RequestJson)

ResultJson = sheet['F2'].value
dict_result = ast.literal_eval(ResultJson)

#request 파싱
arr_request = []

for key, value in dict_request.items():
    if (type(value) == dict):
        for key2, value2 in value.items():
            if key2 in request_replace:
                key2 = request_replace[key2]
            arr_request.append(str(key2) + ':' + str(value2))
    else:
        if key in request_replace:
            key = request_replace[key]
        arr_request.append(str(key) + ':' + str(value))

#result 파싱
arr_result = []

for key, value in dict_result.items():
    if (type(value) == dict):
        for key2, value2 in value.items():
            if key2 in result_replace:
                key2 = result_replace[key2]
            arr_result.append(str(key2) + ':' + str(value2))
    else:
        if key in result_replace:
            key = result_replace[key]
        arr_result.append(str(key) + ':' + str(value))


#신규 시트 생성
book.create_sheet(index = 1, title = 'result')
book.save(filename)

sheet = book.worksheets[1]

w = sheet.cell(row=1, column=1)
x = sheet.cell(row=1, column=2)
y = sheet.cell(row=1, column=3)

w.value= str(api)
x.value= str(arr_request)
y.value= str(arr_result)

book.save(filename)

'''
0. api 내용에 따라 행동 분류
1. request, result 열 데이터 가져오기
2. json 파싱
3. 변수별 내용 정리

https://m.blog.naver.com/hankrah/222057818711
'''