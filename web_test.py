#엑셀 읽어오기
import openpyxl



#data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
wb = openpyxl.load_workbook('C:\\workspace\selenium\qaqa.xlsx', data_only=True)

#시트 이름으로 불러오기
ws = wb['EN']
#셀 주소로 값 출력
print(ws['B4'].value)
#셀 좌표로 값 출력
print(ws.cell(4,6).value)
